"""
tests/test_bulk_mail.py

Тесты для routes/bulk_mail.py:
  - _substitute()        — подстановка плейсхолдеров
  - _rows_to_result()    — преобразование сырых строк в словари
  - /api/bulk/parse      — загрузка и парсинг xlsx/ods файлов
  - /api/bulk/send-test  — тестовая отправка письма
  - /api/bulk/send/start — старт задачи рассылки
  - /api/bulk/send/cancel — отмена задачи

Запуск: pytest tests/test_bulk_mail.py -v
"""

import io
import json
import os
import sys
import tempfile
import threading
import time
import unittest.mock as mock
from unittest.mock import MagicMock, patch, call

# ─── Мокаем Windows-зависимости до импорта app ───────────────────────────────
for mod in ('win32com', 'win32com.client', 'pythoncom', 'pywin32_runtime'):
    if mod not in sys.modules:
        sys.modules[mod] = MagicMock()
sys.modules['win32com'].client = MagicMock()
sys.modules.setdefault('exchangelib', MagicMock())
sys.modules.setdefault('exchangelib.errors', MagicMock())

FAKE_NETWORK = tempfile.mkdtemp(prefix='eb_bm_net_')
FAKE_CACHE   = tempfile.mkdtemp(prefix='eb_bm_cache_')

with mock.patch.dict(os.environ, {'APP_MODE': 'admin'}):
    sys.modules.setdefault('app_admin', MagicMock())
    sys.modules.setdefault('app_user',  MagicMock())
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

    import app as email_app
    email_app.NETWORK_RESOURCES_PATH = FAKE_NETWORK
    email_app.CACHE_DIR              = FAKE_CACHE
    email_app.CACHE_VERSION_FILE     = os.path.join(FAKE_CACHE, 'cache_version.txt')
    email_app.NETWORK_VERSION_FILE   = os.path.join(FAKE_NETWORK, 'version.txt')

import pytest
from routes.bulk_mail import _substitute, _rows_to_result


# ─── Фикстуры ────────────────────────────────────────────────────────────────

@pytest.fixture(scope='session')
def app_instance():
    email_app.app.config['TESTING'] = True
    yield email_app.app


@pytest.fixture
def client(app_instance):
    return app_instance.test_client()


@pytest.fixture(autouse=True)
def reset_paths():
    email_app.NETWORK_RESOURCES_PATH = FAKE_NETWORK
    email_app.CACHE_DIR              = FAKE_CACHE
    yield


# ─── _substitute ─────────────────────────────────────────────────────────────

class TestSubstitute:

    def test_simple_replacement(self):
        result = _substitute('Привет, {{ФИО}}!', {'{{ФИО}}': 'ФИО'}, {'ФИО': 'Иван Иванов'})
        assert result == 'Привет, Иван Иванов!'

    def test_multiple_placeholders(self):
        tpl = '{{ФИО}}, {{Должность}}'
        mapping = {'{{ФИО}}': 'ФИО', '{{Должность}}': 'Должность'}
        row = {'ФИО': 'Мария', 'Должность': 'Менеджер'}
        result = _substitute(tpl, mapping, row)
        assert result == 'Мария, Менеджер'

    def test_html_escaping(self):
        result = _substitute('{{X}}', {'{{X}}': 'X'}, {'X': '<b>bold</b>'})
        assert '&lt;b&gt;' in result
        assert '<b>' not in result

    def test_missing_column_gives_empty(self):
        result = _substitute('{{ФИО}}', {'{{ФИО}}': 'НесуществующаяКолонка'}, {'ФИО': 'Иван'})
        assert result == ''

    def test_missing_key_in_row(self):
        result = _substitute('{{X}}', {'{{X}}': 'X'}, {})
        assert result == ''

    def test_span_wrapper_replaced(self):
        tpl = '<span class="bm-inline-ph" data-ph="{{ФИО}}">{{ФИО}}</span>'
        result = _substitute(tpl, {'{{ФИО}}': 'ФИО'}, {'ФИО': 'Тест'})
        assert 'bm-inline-ph' not in result
        assert 'Тест' in result

    def test_empty_template(self):
        result = _substitute('', {'{{ФИО}}': 'ФИО'}, {'ФИО': 'Иван'})
        assert result == ''

    def test_empty_mapping(self):
        result = _substitute('Текст без плейсхолдеров', {}, {'ФИО': 'Иван'})
        assert result == 'Текст без плейсхолдеров'

    def test_none_value_gives_empty(self):
        result = _substitute('{{X}}', {'{{X}}': 'X'}, {'X': None})
        assert result == ''

    def test_ampersand_escaped(self):
        result = _substitute('{{X}}', {'{{X}}': 'X'}, {'X': 'A&B'})
        assert '&amp;' in result

    def test_quotes_escaped(self):
        result = _substitute('{{X}}', {'{{X}}': 'X'}, {'X': '"quoted"'})
        assert '&quot;' in result


# ─── _rows_to_result ─────────────────────────────────────────────────────────

class TestRowsToResult:

    def test_basic(self):
        rows = [['ФИО', 'Email'], ['Иван', 'ivan@test.ru'], ['Мария', 'maria@test.ru']]
        result = _rows_to_result(rows, ['Sheet1'], header_row=1)
        assert result['headers'] == ['ФИО', 'Email']
        assert result['total'] == 2
        assert result['rows'][0] == {'ФИО': 'Иван', 'Email': 'ivan@test.ru'}

    def test_empty_rows(self):
        result = _rows_to_result([], ['Sheet1'], header_row=1)
        assert result['headers'] == []
        assert result['rows'] == []
        assert result['total'] == 0

    def test_trailing_empty_columns_stripped(self):
        rows = [['ФИО', 'Email', '', ''], ['Иван', 'ivan@test.ru', '', '']]
        result = _rows_to_result(rows, ['Sheet1'], header_row=1)
        assert result['headers'] == ['ФИО', 'Email']

    def test_blank_rows_skipped(self):
        rows = [['ФИО', 'Email'], ['', ''], ['Иван', 'ivan@test.ru']]
        result = _rows_to_result(rows, ['Sheet1'], header_row=1)
        assert result['total'] == 1

    def test_header_row_offset(self):
        rows = [['Пропустить'], ['ФИО', 'Email'], ['Иван', 'ivan@test.ru']]
        result = _rows_to_result(rows, ['Sheet1'], header_row=2)
        assert result['headers'] == ['ФИО', 'Email']
        assert result['total'] == 1

    def test_sheets_preserved(self):
        rows = [['ФИО'], ['Иван']]
        result = _rows_to_result(rows, ['Лист1', 'Лист2'], header_row=1)
        assert result['sheets'] == ['Лист1', 'Лист2']

    def test_short_row_padded(self):
        rows = [['ФИО', 'Email', 'Телефон'], ['Иван']]
        result = _rows_to_result(rows, ['Sheet1'], header_row=1)
        assert result['rows'][0]['Email'] == ''
        assert result['rows'][0]['Телефон'] == ''

    def test_header_whitespace_stripped(self):
        rows = [['  ФИО  ', ' Email '], ['Иван', 'ivan@test.ru']]
        result = _rows_to_result(rows, ['Sheet1'], header_row=1)
        assert 'ФИО' in result['headers']
        assert 'Email' in result['headers']


# ─── /api/bulk/parse ─────────────────────────────────────────────────────────

class TestBulkParse:

    def test_no_file_returns_400(self, client):
        resp = client.post('/api/bulk/parse')
        assert resp.status_code == 400
        assert 'error' in resp.get_json()

    def test_unsupported_format_returns_400(self, client):
        data = {'file': (io.BytesIO(b'data'), 'test.csv')}
        resp = client.post('/api/bulk/parse', data=data, content_type='multipart/form-data')
        assert resp.status_code == 400
        assert 'error' in resp.get_json()

    def test_xlsx_parsed(self, client):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['ФИО', 'Email'])
        ws.append(['Иван', 'ivan@test.ru'])
        ws.append(['Мария', 'maria@test.ru'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        data = {'file': (buf, 'test.xlsx')}
        resp = client.post('/api/bulk/parse', data=data, content_type='multipart/form-data')
        assert resp.status_code == 200
        body = resp.get_json()
        assert body['headers'] == ['ФИО', 'Email']
        assert body['total'] == 2

    def test_xlsx_with_sheet_param(self, client):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Данные'
        ws.append(['Имя', 'Почта'])
        ws.append(['Тест', 'test@test.ru'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        data = {'file': (buf, 'test.xlsx'), 'sheet': 'Данные'}
        resp = client.post('/api/bulk/parse', data=data, content_type='multipart/form-data')
        assert resp.status_code == 200
        body = resp.get_json()
        assert body['headers'] == ['Имя', 'Почта']

    def test_xlsx_header_row_param(self, client):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Пропустить эту строку'])
        ws.append(['ФИО', 'Email'])
        ws.append(['Иван', 'ivan@test.ru'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        data = {'file': (buf, 'test.xlsx'), 'header_row': '2'}
        resp = client.post('/api/bulk/parse', data=data, content_type='multipart/form-data')
        assert resp.status_code == 200
        body = resp.get_json()
        assert body['headers'] == ['ФИО', 'Email']
        assert body['total'] == 1

    def test_corrupt_xlsx_returns_422(self, client):
        data = {'file': (io.BytesIO(b'not-an-xlsx'), 'bad.xlsx')}
        resp = client.post('/api/bulk/parse', data=data, content_type='multipart/form-data')
        assert resp.status_code == 422


# ─── /api/bulk/send-test ─────────────────────────────────────────────────────

class TestBulkSendTest:

    def _mock_creds(self):
        return {
            'server':     'exchange.test.ru',
            'username':   'user@test.ru',
            'password':   'pass',
            'from_email': 'sender@test.ru',
            'auth_type':  'ntlm',
        }

    def test_no_credentials_returns_401(self, client):
        with patch.object(email_app, 'credentials_exist', return_value=False):
            resp = client.post('/api/bulk/send-test',
                               json={'template_html': '<p>test</p>', 'row': {}, 'mapping': {}})
        assert resp.status_code == 401

    def test_success(self, client):
        creds = self._mock_creds()
        mock_account = MagicMock()
        with patch.object(email_app, 'credentials_exist', return_value=True), \
             patch.object(email_app, 'load_credentials', return_value=creds), \
             patch.object(email_app, 'connect_exchange', return_value=mock_account), \
             patch.object(email_app, 'exchange_send_email') as mock_send, \
             patch.object(email_app, 'prepare_html_for_email', return_value='<html/>'):
            resp = client.post('/api/bulk/send-test', json={
                'template_html': '<p>Привет {{ФИО}}</p>',
                'row':     {'ФИО': 'Иван'},
                'mapping': {'{{ФИО}}': 'ФИО'},
                'subject': 'Тест',
            })
        assert resp.status_code == 200
        data = resp.get_json()
        assert data['success'] is True
        assert data['to'] == 'sender@test.ru'
        mock_send.assert_called_once()

    def test_subject_prefixed_with_test(self, client):
        creds = self._mock_creds()
        mock_account = MagicMock()
        captured = {}
        def capture_send(account, subject, body, to, *args, **kwargs):
            captured['subject'] = subject
        with patch.object(email_app, 'credentials_exist', return_value=True), \
             patch.object(email_app, 'load_credentials', return_value=creds), \
             patch.object(email_app, 'connect_exchange', return_value=mock_account), \
             patch.object(email_app, 'exchange_send_email', side_effect=capture_send), \
             patch.object(email_app, 'prepare_html_for_email', return_value='<html/>'):
            client.post('/api/bulk/send-test', json={
                'template_html': '<p>test</p>',
                'row':     {},
                'mapping': {},
                'subject': 'Мой сабж',
            })
        assert captured.get('subject', '').startswith('[Тест]')
        assert 'Мой сабж' in captured.get('subject', '')

    def test_placeholder_substituted_in_body(self, client):
        creds = self._mock_creds()
        mock_account = MagicMock()
        captured = {}
        def capture_prepare(html):
            captured['html'] = html
            return html
        with patch.object(email_app, 'credentials_exist', return_value=True), \
             patch.object(email_app, 'load_credentials', return_value=creds), \
             patch.object(email_app, 'connect_exchange', return_value=mock_account), \
             patch.object(email_app, 'exchange_send_email'), \
             patch.object(email_app, 'prepare_html_for_email', side_effect=capture_prepare):
            client.post('/api/bulk/send-test', json={
                'template_html': 'Привет {{ФИО}}',
                'row':     {'ФИО': 'Иван'},
                'mapping': {'{{ФИО}}': 'ФИО'},
            })
        assert 'Иван' in captured.get('html', '')
        assert '{{ФИО}}' not in captured.get('html', '')

    def test_exchange_error_returns_503(self, client):
        with patch.object(email_app, 'credentials_exist', return_value=True), \
             patch.object(email_app, 'load_credentials', return_value=self._mock_creds()), \
             patch.object(email_app, 'connect_exchange',
                          side_effect=ConnectionError('Сервер недоступен')):
            resp = client.post('/api/bulk/send-test',
                               json={'template_html': '<p/>', 'row': {}, 'mapping': {}})
        assert resp.status_code == 503


# ─── /api/bulk/send/start + stream + cancel ──────────────────────────────────

class TestBulkSendStart:

    def _mock_creds(self):
        return {
            'server':     'exchange.test.ru',
            'username':   'user@test.ru',
            'password':   'pass',
            'from_email': 'sender@test.ru',
            'auth_type':  'ntlm',
        }

    def test_returns_job_id(self, client):
        creds = self._mock_creds()
        mock_account = MagicMock()
        with patch.object(email_app, 'credentials_exist', return_value=True), \
             patch.object(email_app, 'load_credentials', return_value=creds), \
             patch.object(email_app, 'connect_exchange', return_value=mock_account), \
             patch.object(email_app, 'exchange_send_email'), \
             patch.object(email_app, 'prepare_html_for_email', return_value='<html/>'):
            resp = client.post('/api/bulk/send/start', json={
                'template_html': '<p>{{ФИО}}</p>',
                'rows':    [{'ФИО': 'Иван', 'Email': 'ivan@test.ru'}],
                'mapping': {'{{ФИО}}': 'ФИО'},
                'subject': 'Тест',
                'email_column': 'Email',
            })
        assert resp.status_code == 200
        data = resp.get_json()
        assert 'job_id' in data
        assert len(data['job_id']) == 16

    def test_cancel_unknown_job_returns_ok(self, client):
        # cancel всегда возвращает 200 — даже для несуществующих задач
        resp = client.post('/api/bulk/send/cancel/nonexistent_job_id_xyz')
        assert resp.status_code == 200
        assert resp.get_json()['ok'] is True

    def test_stream_unknown_job_returns_404(self, client):
        resp = client.get('/api/bulk/send/stream/nonexistent_job_id_xyz')
        assert resp.status_code == 404

    def test_draft_mode_uses_save_draft(self, client):
        creds = self._mock_creds()
        mock_account = MagicMock()
        with patch.object(email_app, 'credentials_exist', return_value=True), \
             patch.object(email_app, 'load_credentials', return_value=creds), \
             patch.object(email_app, 'connect_exchange', return_value=mock_account), \
             patch.object(email_app, 'exchange_save_draft') as mock_draft, \
             patch.object(email_app, 'exchange_send_email') as mock_send, \
             patch.object(email_app, 'prepare_html_for_email', return_value='<html/>'):
            resp = client.post('/api/bulk/send/start', json={
                'template_html': '<p>test</p>',
                'rows':    [{'ФИО': 'Иван', 'Email': 'ivan@test.ru'}],
                'mapping': {},
                'subject': 'Тест',
                'email_column': 'Email',
                'draft_mode': True,
            })
            assert resp.status_code == 200
            time.sleep(0.3)

        mock_send.assert_not_called()
        mock_draft.assert_called_once()

    def test_no_credentials_job_emits_error_event(self, client):
        with patch.object(email_app, 'credentials_exist', return_value=False):
            resp = client.post('/api/bulk/send/start', json={
                'template_html': '<p>test</p>',
                'rows':    [{'Email': 'ivan@test.ru'}],
                'mapping': {},
                'subject': 'Тест',
                'email_column': 'Email',
            })
        assert resp.status_code == 200
        job_id = resp.get_json()['job_id']

        time.sleep(0.3)
        stream_resp = client.get(f'/api/bulk/send/stream/{job_id}',
                                 headers={'Accept': 'text/event-stream'})
        raw = stream_resp.get_data(as_text=True)
        assert 'error' in raw or 'cancelled' in raw or 'done' in raw
