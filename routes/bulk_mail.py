"""
Flask blueprint: bulk mail — file parsing and sending.

Routes:
    POST /api/bulk/parse              — parse Excel/ODS file
    POST /api/bulk/send-test          — send test email to self
    POST /api/bulk/send/start         — start bulk send job, returns job_id
    GET  /api/bulk/send/stream/<id>   — SSE stream: progress events
    POST /api/bulk/send/cancel/<id>   — cancel running job
"""
from __future__ import annotations

import base64
import datetime
import html as _html
import io
import json
import logging
import mimetypes
import os
import queue
import re
import threading
import uuid

_logger = logging.getLogger(__name__)

from flask import Blueprint, Response, jsonify, request, stream_with_context

import app as _m  # live access to app.py globals (same pattern as exchange.py)

bp = Blueprint('bulk_mail', __name__)

# ── In-memory job store ───────────────────────────────────────────────────────
_jobs: dict = {}
_jobs_lock  = threading.Lock()


def _job_id() -> str:
    return uuid.uuid4().hex[:16]


# ── Placeholder substitution ─────────────────────────────────────────────────

_URL_RE = re.compile(r'^https?://', re.I)
_LINK_COLOR = '#7700ff'


def _substitute(template: str, mapping: dict, row: dict) -> str:
    """Replace {{ColName}} with row values (HTML-escaped).

    Handles both plain ``{{X}}`` text and spans inserted by the {{}} toolbar:
    ``<span class="bm-inline-ph" ...>{{X}}</span>``

    URL values in text content are wrapped in <a> with the brand link colour so
    that all clients (including Evolution on Linux) render them as styled links
    instead of auto-linking them in the default client blue.
    Attribute contexts (href="{{X}}", src="{{X}}") receive the plain value.
    """
    result = template
    for ph, col in mapping.items():
        raw = str(row.get(col) or '')
        value = _html.escape(raw)

        if _URL_RE.match(raw):
            linked = (
                f'<a href="{value}" '
                f'style="color:{_LINK_COLOR}; text-decoration:underline;">'
                f'{value}</a>'
            )
        else:
            linked = value

        # 1. Span wrapper inserted by {{}} toolbar → text context → linked value
        result = re.sub(
            r'<span[^>]+class="[^"]*bm-inline-ph[^"]*"[^>]*>' + re.escape(ph) + r'</span>',
            linked,
            result,
        )

        # 2. Inside href/src/action attributes → plain value (no <a> wrapping)
        result = re.sub(
            r'((?:href|src|action)=")' + re.escape(ph) + r'(")',
            lambda m, v=value: m.group(1) + v + m.group(2),
            result,
        )

        # 3. Any remaining occurrence is text content → linked value
        result = result.replace(ph, linked)

    return result


def _connect(from_email: str = '', channel: str = 'exchange') -> tuple:
    """Load credentials and return (account_or_smtp, creds) or raise.

    Returns:
        channel='exchange' → (exchangelib.Account, creds)
        channel='smtp'     → (smtplib.SMTP, creds)
    """
    path = _m.get_credentials_path()
    if not _m.credentials_exist(path):
        raise ValueError('Учётные данные не настроены')
    creds = _m.load_credentials(path)
    if not creds:
        raise ValueError('Не удалось загрузить учётные данные')

    if channel == 'smtp':
        host     = creds.get('smtp_host', '')
        port     = int(creds.get('smtp_port') or 587)
        username = creds.get('smtp_username', '')
        password = creds.get('smtp_password', '')
        if not host or not username:
            raise ValueError('SMTP не настроен. Заполните настройки SMTP.')
        smtp = _m.connect_smtp(host, port, username, password)
        return smtp, creds

    # Exchange
    effective_from = from_email.strip() or creds.get('from_email', '')
    account = _m.connect_exchange(
        creds['server'], creds['username'], creds['password'],
        effective_from,
        auth_type=creds.get('auth_type', 'ntlm'),
    )
    return account, creds


# ── Parse ─────────────────────────────────────────────────────────────────────

@bp.route('/api/bulk/parse', methods=['POST'])
def api_bulk_parse():
    """Parse an uploaded .xlsx / .xls / .ods file.

    Form fields:
        file        — multipart file upload
        sheet       — sheet name to read (optional, defaults to first sheet)
        header_row  — 1-based row number that contains column headers (default 1)

    Returns JSON:
        {
            "sheets":  ["Sheet1", "Sheet2", ...],
            "headers": ["ФИО", "Email", ...],
            "rows":    [{"ФИО": "...", "Email": "..."}, ...],
            "total":   127
        }
    """
    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'error': 'Файл не передан'}), 400

    sheet_name = request.form.get('sheet') or None
    try:
        header_row = max(1, int(request.form.get('header_row', 1)))
    except (ValueError, TypeError):
        header_row = 1

    fname = file.filename.lower()
    try:
        file_bytes = file.read()
        if fname.endswith(('.xlsx', '.xlsm')):
            result = _parse_xlsx(file_bytes, sheet_name, header_row)
        elif fname.endswith('.xls'):
            result = _parse_xls(file_bytes, sheet_name, header_row)
        elif fname.endswith('.ods'):
            result = _parse_ods(file_bytes, sheet_name, header_row)
        elif fname.endswith('.csv'):
            result = _parse_csv(file_bytes, header_row)
        else:
            return jsonify({'error': (
                f'Неподдерживаемый формат: «{file.filename}». '
                'Разрешены: .xlsx, .xls, .ods, .csv'
            )}), 400
    except ImportError as e:
        return jsonify({'error': str(e)}), 501
    except Exception as e:
        return jsonify({'error': f'Ошибка чтения файла: {e}'}), 422

    return jsonify(result)


# ── Parsers ───────────────────────────────────────────────────────────────────

def _parse_xlsx(file_bytes: bytes, sheet_name: str | None, header_row: int) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheets = wb.sheetnames
    ws = wb[sheet_name] if (sheet_name and sheet_name in wb) else wb.active
    rows_raw = [
        ['' if cell is None else str(cell) for cell in row]
        for row in ws.iter_rows(values_only=True)
    ]
    wb.close()
    return _rows_to_result(rows_raw, sheets, header_row)


def _parse_xls(file_bytes: bytes, sheet_name: str | None, header_row: int) -> dict:
    try:
        import xlrd
    except ImportError:
        raise ImportError('Для .xls-файлов установите: pip install "xlrd<2.0"')
    wb = xlrd.open_workbook(file_contents=file_bytes)
    sheets = wb.sheet_names()
    ws = wb.sheet_by_name(sheet_name) if (sheet_name and sheet_name in sheets) else wb.sheet_by_index(0)
    rows_raw = [
        ['' if v is None else str(v) for v in ws.row_values(i)]
        for i in range(ws.nrows)
    ]
    return _rows_to_result(rows_raw, sheets, header_row)


def _parse_ods(file_bytes: bytes, sheet_name: str | None, header_row: int) -> dict:
    try:
        from odf.opendocument import load as ods_load
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
    except ImportError:
        raise ImportError('Для .ods-файлов установите: pip install odfpy')

    doc = ods_load(io.BytesIO(file_bytes))
    tables = doc.spreadsheet.getElementsByType(Table)
    sheets = [t.getAttribute('name') for t in tables]

    target = next(
        (t for t in tables if t.getAttribute('name') == sheet_name),
        tables[0] if tables else None,
    )
    if target is None:
        raise ValueError('Лист не найден')

    rows_raw: list[list[str]] = []
    for tr in target.getElementsByType(TableRow):
        row: list[str] = []
        for tc in tr.getElementsByType(TableCell):
            repeat = int(tc.getAttribute('numbercolumnsrepeated') or 1)
            ps = tc.getElementsByType(P)
            val = ''.join(str(n) for n in ps) if ps else ''
            row.extend([val] * repeat)
        # strip trailing empty cells added by repeat
        while row and not row[-1].strip():
            row.pop()
        rows_raw.append(row)

    return _rows_to_result(rows_raw, sheets, header_row)


def _parse_csv(file_bytes: bytes, header_row: int) -> dict:
    import csv, io as _io
    # Try common encodings (BOM-aware utf-8-sig first, then cp1251 for Windows files)
    for enc in ('utf-8-sig', 'utf-8', 'cp1251', 'latin-1'):
        try:
            text = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = file_bytes.decode('utf-8', errors='replace')

    # Auto-detect delimiter from the first 8 KB
    sample = text[:8192]
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(sample, delimiters=',;\t|')
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ','  # fallback

    reader = csv.reader(_io.StringIO(text), delimiter=delimiter)
    rows_raw = [row for row in reader]
    return _rows_to_result(rows_raw, [], header_row)


# ── Shared conversion ─────────────────────────────────────────────────────────

def _rows_to_result(rows_raw: list[list[str]], sheets: list[str], header_row: int) -> dict:
    """Convert a list-of-lists to the standard API response dict."""
    if not rows_raw:
        return {'sheets': sheets, 'headers': [], 'rows': [], 'total': 0}

    h_idx = min(header_row - 1, len(rows_raw) - 1)
    header_cells = rows_raw[h_idx]

    # Drop trailing empty columns
    while header_cells and not str(header_cells[-1]).strip():
        header_cells = header_cells[:-1]

    headers = [str(c).strip() for c in header_cells]
    col_count = len(headers)
    if col_count == 0:
        return {'sheets': sheets, 'headers': [], 'rows': [], 'total': 0}

    data_rows: list[dict] = []
    for raw in rows_raw[h_idx + 1:]:
        padded = (list(raw) + [''] * col_count)[:col_count]
        row_dict = {headers[i]: str(padded[i]).strip() for i in range(col_count)}
        # Skip completely blank rows
        if any(v for v in row_dict.values()):
            data_rows.append(row_dict)

    return {
        'sheets': sheets,
        'headers': headers,
        'rows': data_rows,
        'total': len(data_rows),
    }


# ── Send test ─────────────────────────────────────────────────────────────────

@bp.route('/api/bulk/send-test', methods=['POST'])
def api_bulk_send_test():
    """Send one test email to the logged-in user substituting the given row."""
    data = request.get_json(silent=True) or {}
    template_html = data.get('template_html', '')
    row           = data.get('row', {})
    mapping       = data.get('mapping', {})
    subject_tpl   = data.get('subject', 'Тестовое письмо')

    try:
        from_email   = data.get('from_email', '').strip()
        channel      = str(data.get('channel')    or 'exchange').lower()
        importance   = str(data.get('importance') or 'normal').lower()
        read_receipt = bool(data.get('read_receipt'))
        conn, creds = _connect(from_email, channel)
        subject = '[Тест] ' + _substitute(subject_tpl, mapping, row)
        body    = _m.prepare_html_for_email(_substitute(template_html, mapping, row))

        if channel == 'smtp':
            smtp_from = from_email or creds.get('smtp_from_email', '')
            to        = smtp_from
            _m.smtp_send_email(conn, smtp_from, subject, body, [smtp_from],
                               importance=importance, read_receipt=read_receipt)
            try: conn.quit()
            except Exception: pass
        else:
            to = from_email or creds.get('from_email', '')
            _m.exchange_send_email(conn, subject, body, _m.parse_recipients([to]),
                                   importance=importance, read_receipt=read_receipt)

        return jsonify({'success': True, 'to': to})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 401
    except ConnectionError as e:
        return jsonify({'success': False, 'error': str(e)}), 503
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ── Bulk send ─────────────────────────────────────────────────────────────────

@bp.route('/api/bulk/send/start', methods=['POST'])
def api_bulk_send_start():
    """Queue a bulk-send job. Returns ``{job_id}`` immediately."""
    data   = request.get_json(silent=True) or {}
    jid    = _job_id()
    cancel = threading.Event()
    q: queue.Queue = queue.Queue()

    with _jobs_lock:
        _jobs[jid] = {'queue': q, 'cancel': cancel}

    threading.Thread(
        target=_run_bulk_send,
        args=(jid, data, q, cancel),
        daemon=True,
    ).start()

    return jsonify({'job_id': jid})


@bp.route('/api/bulk/send/stream/<job_id>', methods=['GET'])
def api_bulk_send_stream(job_id):
    """SSE stream for a running bulk-send job."""
    with _jobs_lock:
        job = _jobs.get(job_id)
    if not job:
        return jsonify({'error': 'Задание не найдено'}), 404

    def generate():
        q = job['queue']
        while True:
            try:
                event = q.get(timeout=30)
                yield f'data: {json.dumps(event, ensure_ascii=False)}\n\n'
                if event.get('type') in ('done', 'cancelled', 'error'):
                    with _jobs_lock:
                        _jobs.pop(job_id, None)
                    break
            except queue.Empty:
                yield 'data: {"type":"heartbeat"}\n\n'

    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'},
    )


@bp.route('/api/bulk/send/cancel/<job_id>', methods=['POST'])
def api_bulk_send_cancel(job_id):
    """Signal a running job to stop."""
    with _jobs_lock:
        job = _jobs.get(job_id)
    if job:
        job['cancel'].set()
    return jsonify({'ok': True})


# ── Background worker ─────────────────────────────────────────────────────────

def _resolve_attachment(folder: str, filename_tpl: str, mapping: dict, row: dict) -> str | None:
    """Return the full path to an attachment file, or None if not found/not configured."""
    if not folder or not filename_tpl:
        return None
    # Substitute placeholders in filename template (plain text, no HTML escape)
    fname = filename_tpl
    for ph, col in mapping.items():
        fname = fname.replace(ph, str(row.get(col) or ''))
    # Sanitize: remove characters that are illegal in filenames
    fname = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', fname)
    path = os.path.join(folder, fname)
    return path if os.path.isfile(path) else None


def _run_bulk_send(job_id: str, data: dict, q: queue.Queue, cancel: threading.Event):
    rows          = data.get('rows', [])
    mapping       = data.get('mapping', {})
    template_html = data.get('template_html', '')
    subject_tpl   = data.get('subject', '')
    email_col     = data.get('email_column', 'Email')
    cc_raw        = data.get('cc',  '')
    bcc_raw       = data.get('bcc', '')
    skip_no_email = data.get('skip_no_email', True)
    is_draft      = data.get('draft_mode', False)
    stop_on_error = data.get('stop_on_error', False)
    delay         = float(data.get('delay', 0))
    channel        = str(data.get('channel')     or 'exchange').lower()
    from_email_req = str(data.get('from_email')  or '').strip()
    importance     = str(data.get('importance')  or 'normal').lower()
    read_receipt   = bool(data.get('read_receipt'))

    send_at = None
    send_at_raw = data.get('send_at', '')
    if send_at_raw:
        try:
            send_at = _m.parse_datetime(send_at_raw)
            if send_at <= datetime.datetime.now():
                q.put({'type': 'error',
                       'message': 'Дата отложенной отправки должна быть в будущем'})
                return
        except (ValueError, TypeError) as e:
            _logger.warning('bulk_send: invalid send_at %r: %s', send_at_raw, e)

    timezone = float(data.get('timezone') if data.get('timezone') is not None else 3.0)
    total    = len(rows)

    # Attachment settings
    attach_enabled  = data.get('attach_enabled', False)
    attach_folder   = data.get('attach_folder', '').strip()
    attach_tpl      = data.get('attach_template', '').strip()
    attach_missing  = data.get('attach_missing', 'send')   # send | skip | stop

    # Prepare static CC / BCC once
    cc  = _m.parse_recipients([cc_raw])  if cc_raw  else []
    bcc = _m.parse_recipients([bcc_raw]) if bcc_raw else []

    # Connect once for all rows
    try:
        conn, creds = _connect(from_email_req, channel)
    except Exception as e:
        q.put({'type': 'error', 'message': str(e)})
        return

    # Determine effective from_email and SMTP-specific settings
    if channel == 'smtp':
        smtp_from      = from_email_req or creds.get('smtp_from_email', '')
        smtp_imap      = creds.get('smtp_imap_enabled', False)
        smtp_imap_host = creds.get('smtp_imap_host', '')
        smtp_imap_port = int(creds.get('smtp_imap_port') or 993)
        smtp_delay_on  = creds.get('smtp_delay_enabled', False)
        smtp_delay_sec = float(creds.get('smtp_delay_seconds') or 1)
        # Открываем IMAP если нужно
        imap_conn = None
        if smtp_imap and smtp_imap_host:
            try:
                imap_conn = _m.connect_imap(
                    smtp_imap_host, smtp_imap_port,
                    creds.get('smtp_username', ''),
                    creds.get('smtp_password', ''),
                )
            except Exception as e:
                _logger.warning('bulk_send: IMAP connect failed: %s', e)
    else:
        account = conn

    sent = skipped = errors = 0

    # Helper to resolve the display name from mapping
    fio_col = mapping.get('{{ФИО}}', 'ФИО')

    for i, row in enumerate(rows):
        if cancel.is_set():
            q.put({'type': 'cancelled', 'sent': sent, 'skipped': skipped, 'errors': errors})
            return

        email = str(row.get(email_col) or '').strip()
        name  = str(row.get(fio_col) or '').strip()

        if skip_no_email and not email:
            skipped += 1
            q.put({'type': 'progress', 'index': i, 'total': total,
                   'name': name, 'email': '(нет email)', 'status': 'skip',
                   'comment': 'Пропущено'})
            continue

        try:
            subject = _substitute(subject_tpl, mapping, row)
            body    = _m.prepare_html_for_email(_substitute(template_html, mapping, row))
            to      = _m.parse_recipients([email])

            # Resolve attachment
            attachments = []
            if attach_enabled and attach_folder and attach_tpl:
                attach_path = _resolve_attachment(attach_folder, attach_tpl, mapping, row)
                if attach_path:
                    # exchange_send_email expects: {'name': str, 'content': base64str, 'mime_type': str}
                    try:
                        with open(attach_path, 'rb') as fh:
                            raw = fh.read()
                        mime_type, _ = mimetypes.guess_type(attach_path)
                        attachments = [{
                            'name':      os.path.basename(attach_path),
                            'content':   base64.b64encode(raw).decode('utf-8'),
                            'mime_type': mime_type or 'application/octet-stream',
                        }]
                    except Exception as e:
                        _logger.warning('bulk_send: cannot read attachment %s: %s', attach_path, e)
                        attachments = []
                elif attach_missing == 'skip':
                    skipped += 1
                    q.put({'type': 'progress', 'index': i, 'total': total,
                           'name': name, 'email': email, 'status': 'skip',
                           'comment': 'Пропущено: вложение не найдено'})
                    continue
                elif attach_missing == 'stop':
                    q.put({'type': 'cancelled', 'sent': sent, 'skipped': skipped, 'errors': errors,
                           'message': f'Остановлено: вложение не найдено для {name}'})
                    return
                # else 'send': proceed without attachment

            if channel == 'smtp':
                # Вложения для SMTP: bytes вместо base64-строки
                smtp_atts = []
                for att in attachments:
                    content = att.get('content', '')
                    raw_bytes = (base64.b64decode(content)
                                 if isinstance(content, str) else content)
                    smtp_atts.append({
                        'name': att['name'],
                        'content': raw_bytes,
                        'mime_type': att.get('mime_type', 'application/octet-stream'),
                    })
                raw_msg = _m.smtp_send_email(
                    conn, smtp_from, subject, body,
                    [email], cc_raw.split(',') if cc_raw else [],
                    bcc_raw.split(',') if bcc_raw else [],
                    attachments=smtp_atts,
                    importance=importance,
                    read_receipt=read_receipt,
                )
                if imap_conn and raw_msg:
                    try:
                        _m.imap_save_sent(imap_conn, raw_msg)
                    except Exception as e:
                        _logger.warning('bulk_send: imap_save_sent failed: %s', e)
                if smtp_delay_on and smtp_delay_sec > 0:
                    import time as _time
                    _time.sleep(smtp_delay_sec)
            elif is_draft:
                _m.exchange_save_draft(account, subject, body, to, cc, bcc,
                                       attachments=attachments)
            else:
                _m.exchange_send_email(account, subject, body, to, cc, bcc,
                                       attachments=attachments,
                                       send_at=send_at,
                                       timezone=timezone,
                                       importance=importance,
                                       read_receipt=read_receipt)

            sent += 1
            comment = ('Сохранён черновик' if is_draft
                       else f'Запланировано на {send_at.strftime("%d.%m.%Y %H:%M")}' if send_at
                       else 'Отправлено')
            q.put({'type': 'progress', 'index': i, 'total': total,
                   'name': name, 'email': email, 'status': 'sent',
                   'comment': comment})

        except Exception as e:
            errors += 1
            q.put({'type': 'progress', 'index': i, 'total': total,
                   'name': name, 'email': email, 'status': 'error',
                   'comment': str(e)})
            if stop_on_error:
                q.put({'type': 'cancelled', 'sent': sent, 'skipped': skipped, 'errors': errors})
                return

        if delay > 0:
            cancel.wait(timeout=delay)

    # Закрыть соединения
    if channel == 'smtp':
        try: conn.quit()
        except Exception: pass
        if imap_conn:
            try: imap_conn.logout()
            except Exception: pass

    if not is_draft and sent > 0:
        try:
            _m.log_send(sent)
        except Exception:
            _logger.warning('analytics log_send failed', exc_info=True)

    q.put({'type': 'done', 'sent': sent, 'skipped': skipped, 'errors': errors})
